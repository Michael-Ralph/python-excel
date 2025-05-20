from openpyxl import Workbook
from openpyxl import load_workbook

workbook = load_workbook(filename="sample.xlsx")
workbook.sheetnames
#print(workbook.sheetnames)

sheet = workbook.active
sheet
#print(sheet)

sheet.title
#print(sheet.title)

#Method 1 to read cell values
#print(sheet["A1"])
#print(sheet["A1"].value)
#print(sheet["F10"].value)

#Method 2 to read cell values
#print(sheet.cell(row=10, column=6))
# #print(sheet.cell(row=10, column=6).value)

#for value in sheet.iter_rows(min_row=1,
#                             max_row=2,
#                             min_col=1,
#                             max_col=3,
#                             values_only=True):
#    print(value)

#for row in sheet.rows:
#    print(row)

#for value in sheet.iter_rows(min_row=1,
#                             max_row=1,
#                             values_only=True):
#    print(value)

for value in sheet.iter_rows(min_row=1,
                            min_col=4,
                            max_col=7,
                            values_only=True):
        print(value)